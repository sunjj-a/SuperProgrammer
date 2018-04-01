#-*- encoding:gbk -*-

from openpyxl import load_workbook
from openpyxl.styles import Alignment
import urllib, urllib2, os
import ssl
import json
import ConfigParser

import sys
reload(sys)
sys.setdefaultencoding('utf8')

#outputInfo
def outputHintInfo(info, wrap = None):
    if wrap is None:
        print info
    else:
        print (info),

#read userinfo
def readUserInfo(fileName):
    workBook = load_workbook(fileName)
    curSheet = workBook.get_sheet_by_name("Sheet1")

    userInfos = []
    for nIndex in range(curSheet.max_row):
        userName = curSheet["B%d" % (nIndex + 1)].value
        userCard = curSheet["C%d" % (nIndex + 1)].value
        if  userName and userCard:
            userinfo = (userName, userCard)
            userInfos.append(userinfo)
    return userInfos



#save userInfo
def saveUserInfo(fileName, checkInfos):
    workBook = load_workbook(fileName)
    curSheet = workBook.get_sheet_by_name("Sheet1")

    for nIndex in range(len(checkInfos)):
        curSheet["A%d" % (nIndex + 1)].alignment = Alignment(horizontal="left")
        curSheet["A%d" % (nIndex + 1)].value = checkInfos[nIndex]
    workBook.save(fileName)



#check userInfo
def checkUserInfo(appCode, userInfos):
    checkInfos = []
    for user in userInfos:
        bodyData = {}
        bodyData['realName'] = user[0]
        bodyData['cardNo'] = user[1]

        url = "https://1.api.apistore.cn/idcard3"
        postData = urllib.urlencode(bodyData)
        request = urllib2.Request(url, postData)
        request.add_header('Authorization', 'APPCODE ' + appCode)
        request.add_header('Content-Type', 'application/x-www-form-urlencoded; charset=UTF-8')

        context = ssl.create_default_context()
        context.check_hostname = False
        context.verify_mode = ssl.CERT_NONE

        response = urllib2.urlopen(request, context=context)
        checkResult =  response.read()
        if checkResult:
            userJson = json.loads(checkResult, encoding="UTF-8")
            checkInfos.append(userJson["reason"])
        else:
            checkInfos.append("")
    return checkInfos



#process check all user info
def processUserInfo(appCode, fileName):
    outputHintInfo(u"加载用户信息...")
    userInfos = readUserInfo(fileName)
    outputHintInfo(u"验证用户信息...")
    checkInfos = checkUserInfo(appCode, userInfos)
    outputHintInfo(u"保存验证信息...")
    saveUserInfo(fileName, checkInfos)
    outputHintInfo(u"---------------")


#process core
def processCore():
    config = ConfigParser.ConfigParser()
    config.read("UserInfo.ini")
    appCode = config.get("section", "AppCode")

    for (root, dirs, files) in os.walk("UserInfo"):
        for fileName in files:
            if fileName.endswith(".xlsx"):
                outputHintInfo(u"处理文件:", False)
                outputHintInfo(fileName, False)
                outputHintInfo(u"中...")
                fileName = root + "\\" + fileName
                processUserInfo(appCode, fileName)



if __name__ == "__main__":
    processCore()
    raw_input((u"用户身份验证结束^_^\n按任意键退出...").encode("gbk"))